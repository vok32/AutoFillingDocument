import re
import os
import sys
import subprocess
import openpyxl as op
from tkinter import Tk, Label, Button, Listbox, Scrollbar, filedialog, messagebox, StringVar, Frame, LEFT, Text
from docx2txt import process
from docxtpl import DocxTemplate
from docx2pdf import convert
import random
import string
import pyautogui

# Глобальные переменные для хранения путей к файлам и папкам
TEMPLATE_PATH = ""
EXCEL_PATH = ""
SAVE_PATH = ""

# Функция для закрытия программы
def close_program():
    exit()

# Функция для перезапуска программы
def restart_program():
    python = sys.executable
    os.execl(python, python, *sys.argv)

# Функция удаления файла с названием заголовка
def delete_files_with_pattern(directory, pattern):
    for filename in os.listdir(directory):
        if pattern in filename:
            os.remove(os.path.join(directory, filename))

# Функция для генерации уникального суффикса
def generate_unique_suffix(file_name):
    base_name, ext = os.path.splitext(file_name)
    suffix_pattern = re.compile(r'_new_(\d+)')
    existing_suffixes = set()
    
    # Поиск существующих суффиксов в именах файлов
    for file in os.listdir(SAVE_PATH):
        if file.startswith(f"шаблон-{base_name}_new_") and file.endswith(ext):
            match = suffix_pattern.search(file)
            if match:
                existing_suffixes.add(int(match.group(1)))

    # Выбор следующего доступного суффикса
    suffix_num = 1
    while suffix_num in existing_suffixes:
        suffix_num += 1
        
    new_suffix = f"_new_{suffix_num}"
    return base_name + new_suffix + ext

# Функция для парсинга переменных из шаблона Word
def parse_template(template_path):
    text = process(template_path)
    variables = set(re.findall(r'\{\{(.+?)\}\}', text))  # Используем множество для уникальных переменных
    return variables

# Функция для выбора файла шаблона Word
def select_template_file():
    global TEMPLATE_PATH
    global template_file_label

    selected_path = filedialog.askopenfilename(title="Выберите файл шаблона Word", filetypes=[("Word files", "*.docx")])
    if selected_path:
        TEMPLATE_PATH = selected_path
        template_variables = parse_template(TEMPLATE_PATH)
        if not template_variables:
            messagebox.showerror("Ошибка", "В выбранном шаблоне нет переменных. Пожалуйста, выберите другой шаблон.")
            TEMPLATE_PATH = ""  # Сбрасываем путь к файлу шаблона
            template_file_label.config(text="")  # Очищаем метку с именем файла шаблона
            description_label.config(text="В выбранном шаблоне не найдены переменные. Пожалуйста, выберите другой шаблон.")
        else:
            template_file_label.config(text=os.path.basename(TEMPLATE_PATH))
            description_label.config(text="Это программа для автозаполнения шаблонов документов Word при помощи таблицы Excel")
    else:
        messagebox.showerror("Ошибка", "Файл шаблона не выбран. Пожалуйста, выберите файл шаблона.")

# Функция для выбора файла Excel с данными
def select_excel_file():
    global EXCEL_PATH
    EXCEL_PATH = filedialog.askopenfilename(title="Выберите файл с данными Excel", filetypes=[("Excel files", "*.xlsx")])
    excel_file_label.config(text=os.path.basename(EXCEL_PATH))

# Функция для выбора папки для сохранения собранных файлов
def select_save_folder():
    global SAVE_PATH
    SAVE_PATH = filedialog.askdirectory(title="Выберите путь сохранения собранных файлов")
    save_folder_label.config(text=SAVE_PATH)

# Функция для открытия папки с собранными файлами
def open_folder():
    if os.path.exists(SAVE_PATH):
        os.startfile(SAVE_PATH)
    else:
        messagebox.showinfo("Папка не найдена", "Папка с собранными файлами не существует или не была выбрана.")

# Функция для сравнения заголовков Excel и переменных шаблона
def compare_headers_and_variables(header_row, template_variables):
    excel_headers_set = set(header_row.keys())
    template_variables_set = set(template_variables)

    if excel_headers_set != template_variables_set:
        missing_in_excel = excel_headers_set - template_variables_set
        missing_in_template = template_variables_set - excel_headers_set
        message = ""
        if missing_in_excel:
            message += "Заголовки в файле Excel, но отсутствующие в шаблоне:\n"
            for header in missing_in_excel:
                message += f"- {header}\n"
            message += "\n"
        if missing_in_template:
            message += "Переменные в шаблоне, но отсутствующие в заголовках файла Excel:\n"
            for var in missing_in_template:
                message += f"- {var}\n"
        return message
    else:
        return "Заголовки в файле Excel совпадают с переменными в шаблоне."

# Функция для отображения интерфейса выбора заголовков и переменных
def show_header_and_variable_selection_ui(root, header_row, template_variables):
    clear_window(root)
    root.geometry("900x350")  # Установка размера окна

    excel_frame = Frame(root)
    excel_frame.grid(row=0, column=0, padx=10, pady=5)
    
    template_frame = Frame(root)
    template_frame.grid(row=0, column=1, padx=10, pady=5)

    # Создаем метки для заголовков и переменных
    excel_header_label = Label(excel_frame, text="Список заголовков в файле Excel:", font=("Arial", 10, "bold"))
    excel_header_label.grid(row=0, column=0, sticky='w')

    template_header_label = Label(template_frame, text="Переменные в шаблоне Word:", font=("Arial", 10, "bold"))
    template_header_label.grid(row=0, column=0, sticky='w')

    # Сортировка заголовков Excel по алфавиту
    sorted_excel_headers = sorted(header_row.keys())
    excel_headers_text = Text(excel_frame, wrap="word", height=10, width=40)
    excel_headers_text.grid(row=1, column=0, sticky='w')

    for index, column_name in enumerate(sorted_excel_headers, start=1):
        excel_headers_text.insert("end", f"{index}. {column_name}\n")

    # Сортировка переменных шаблона Word по алфавиту
    sorted_template_variables = sorted(template_variables)
    template_variables_text = Text(template_frame, wrap="word", height=10, width=40)
    template_variables_text.grid(row=1, column=0, sticky='w')

    for index, var in enumerate(sorted_template_variables, start=1):
        template_variables_text.insert("end", f"{index}. {var}\n")

    # Добавляем скроллбары
    excel_scrollbar = Scrollbar(excel_frame, command=excel_headers_text.yview)
    excel_scrollbar.grid(row=1, column=1, sticky='ns')
    excel_headers_text.config(yscrollcommand=excel_scrollbar.set)

    template_scrollbar = Scrollbar(template_frame, command=template_variables_text.yview)
    template_scrollbar.grid(row=1, column=1, sticky='ns')
    template_variables_text.config(yscrollcommand=template_scrollbar.set)

    continue_button = Button(root, text="Продолжить", command=lambda: show_differences_ui(root, header_row, template_variables))
    continue_button.grid(row=2, column=0, columnspan=2, pady=5)

    close_button = Button(root, text="Закрыть программу", command=close_program)
    close_button.grid(row=4, column=0, columnspan=2, pady=5)

# Функция для отображения интерфейса различий между заголовками и переменными
def show_differences_ui(root, header_row, template_variables):
    clear_window(root)
    root.geometry("900x350")  # Установка размера окна

    differences_message = compare_headers_and_variables(header_row, template_variables)

    differences_label = Label(root, justify="center")  # Выравнивание по центру
    differences_label.place(relx=0.5, rely=0.4, anchor="center")  # Установка по центру окна и немного выше середины

    # Разделяем сообщение на части и применяем стили к каждой части
    parts = re.split(r'(Заголовки в файле Excel, но отсутствующие в шаблоне:|Переменные в шаблоне, но отсутствующие в заголовках файла Excel:)', differences_message)

    for part in parts:
        if part == "Заголовки в файле Excel, но отсутствующие в шаблоне:":
            label = Label(differences_label, text=part, justify="left", font=("Arial", 12, "bold"), fg="blue")  # Синий цвет
        elif part == "Переменные в шаблоне, но отсутствующие в заголовках файла Excel:":
            label = Label(differences_label, text=part, justify="left", font=("Arial", 12, "bold"), fg="red")
        elif part == "Заголовки в файле Excel совпадают с переменными в шаблоне.":
            label = Label(differences_label, text=part, justify="center", font=("Arial", 12, "bold"), fg="green")  # Зелёный цвет
        else:
            # Для перечисления переменных увеличим размер шрифта
            label = Label(differences_label, text=part, justify="center", font=("Arial", 10))  

        label.pack(anchor="center")  # Пакет по центру

    replace_button = Button(root, text="Продолжить и заменить", command=lambda: select_column(root, header_row, template_variables))
    replace_button.grid(row=1, column=0, columnspan=2, pady=5)  # Используем grid() для управления расположением кнопок

    close_button = Button(root, text="Закрыть программу", command=close_program)
    close_button.grid(row=2, column=0, columnspan=2, pady=5)  # Используем grid() для управления расположением кнопок

    root.grid_rowconfigure(0, weight=1)  # Равномерное распределение по строкам
    root.grid_columnconfigure(0, weight=1)  # Равномерное распределение по столбцам


# Функция для установки ширины метки
def set_label_width(label, max_width):
    label.config(width=max_width)

# Функция для выбора столбца Excel
def select_column(root, header_row, template_variables):
    clear_window(root)

    wb_read = op.load_workbook(filename=EXCEL_PATH, data_only=True)
    sheet_read = wb_read.active

    akt_list = [list(row) for row in sheet_read.iter_rows(values_only=True)]

    def select_column_callback():
        selected_index = listbox.curselection()
        if selected_index:
            selected_column_name = listbox.get(selected_index[0])
            selected_column = header_row[selected_column_name]
            choice_root = Tk()
            choice_root.title("Выбор формата файла")

            def docx_pdf():
                choice_root.destroy()
                create_doc(root, akt_list, header_row, selected_column, selected_column_name, convert_to_pdf=True, delete_docx=False)
                show_success_or_report_window(root)

            def docx_only():
                choice_root.destroy()
                create_doc(root, akt_list, header_row, selected_column, selected_column_name, convert_to_pdf=False)
                show_success_or_report_window(root)

            def pdf_only():
                choice_root.destroy()
                create_doc(root, akt_list, header_row, selected_column, selected_column_name, convert_to_pdf=True, delete_docx=True)
                show_success_or_report_window(root)

            pdf_button = Button(choice_root, text="DOCX and PDF", command=docx_pdf)
            pdf_button.grid(row=0, column=0, padx=10, pady=5)

            docx_button = Button(choice_root, text="DOCX Only", command=docx_only)
            docx_button.grid(row=0, column=1, padx=10, pady=5)

            pdf_only_button = Button(choice_root, text="PDF Only", command=pdf_only)
            pdf_only_button.grid(row=1, column=0, columnspan=2, padx=10, pady=5)

            choice_root.mainloop()
        else:
            messagebox.showinfo("Ошибка", "Не выбрано название для файла. Пожалуйста, выберите название.")

    selection_label = Label(root, text="Выберите название для файлов:")
    selection_label.grid(row=0, column=0, sticky="w")

    root.columnconfigure(0, weight=1)  # Растягиваем первый столбец корневого виджета

    listbox = Listbox(root, selectmode="SINGLE")
    listbox.grid(row=1, column=0, sticky="ew", padx=(10, 0), pady=(20, 10))  # Увеличиваем отступы по горизонтали и вертикали


    scrollbar = Scrollbar(root, orient="vertical")
    scrollbar.config(command=listbox.yview)
    scrollbar.grid(row=1, column=1, sticky="ns")

    listbox.config(yscrollcommand=scrollbar.set)
    for column_name in header_row:
        listbox.insert("end", column_name)

    select_button = Button(root, text="Выбрать", command=select_column_callback)
    select_button.grid(row=2, column=0, pady=5)

    close_button = Button(root, text="Закрыть программу", command=close_program)
    close_button.grid(row=4, column=0, pady=5)

# Функция для отображения окна успешного выполнения или отчёта об ошибке
def show_success_or_report_window(root):
    clear_window(root)
    
    success_label = Label(root, text="Замена прошла успешно", justify="left", fg="green", font=("Arial", 12))
    success_label.grid(row=0, column=0, pady=5)

    open_folder_button = Button(root, text="Открыть папку с собранными файлами", command=open_folder)
    open_folder_button.grid(row=1, column=0, pady=5)

    report_label = Label(root, text="Вы хотите отправить разработчику отчёт об ошибке в приложении?", justify="left")
    report_label.grid(row=2, column=0, pady=5)

    yes_button = Button(root, text="Да, отправить", command=lambda: messagebox.showinfo("Ну и ябеда!", "Ну и ябеда!"))
    yes_button.grid(row=3, column=0, pady=5)

    buttons_frame = Frame(root)
    buttons_frame.grid(row=4, column=0, pady=5)

    restart_button = Button(buttons_frame, text="Запустить программу снова", command=restart_program)
    restart_button.pack(side=LEFT, padx=5)

    close_button = Button(buttons_frame, text="Закрыть программу", command=close_program)
    close_button.pack(side=LEFT, padx=5)

# Функция для очистки основного окна
def clear_window(root):
    for widget in root.winfo_children():
        widget.destroy()

# Функция для создания документа
def create_doc(root, akt_list, header_row, column_index, column_name, convert_to_pdf=True, delete_docx=True):
    header_dict = {name: index for index, name in enumerate(header_row)}
    
    # Проверяем, существуют ли файлы с такими же именами
    existing_files = []
    unique_file_names = {}  # Initialize unique_file_names outside of the conditional block
    for row_data in akt_list:
        context = {}
        for variable, column in header_row.items():
            index = header_dict.get(variable)
            if index is not None:
                context[variable] = row_data[index]
            else:
                context[variable] = ''

        first_row_header = list(header_row.keys())[0]
        file_name = re.sub(r'[\\/*?:"<>|]', '_', str(context[column_name]))

        # Проверяем, существуют ли файлы с такими же именами
        docx_file_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.docx")
        pdf_file_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.pdf")

        if os.path.exists(docx_file_path) or os.path.exists(pdf_file_path):
            existing_files.append((file_name, docx_file_path, pdf_file_path))
            # Check for unique suffixes
            if os.path.exists(docx_file_path):
                unique_suffix = re.findall(r'_(\d+)$', file_name)
                if unique_suffix:
                    unique_file_names.setdefault(file_name.replace(f"_{unique_suffix[0]}", ""), []).append(f"_{unique_suffix[0]}")

    if existing_files:
        existing_files_info = []
        for name, docx_path, pdf_path in existing_files:
            files_info = f'{name}.docx'
            if os.path.exists(pdf_path):
                files_info += f' и {name}.pdf'
            existing_files_info.append(files_info)
            # Check for files with unique suffixes
            for file_suffix in unique_file_names.get(name, []):
                unique_docx_path = os.path.join(SAVE_PATH, f"шаблон-{name}{file_suffix}.docx")
                unique_pdf_path = os.path.join(SAVE_PATH, f"шаблон-{name}{file_suffix}.pdf")
                if os.path.exists(unique_docx_path):
                    files_info = f'{name}{file_suffix}.docx'
                    if os.path.exists(unique_pdf_path):
                        files_info += f' и {name}{file_suffix}.pdf'
                    existing_files_info.append(files_info)

        replace = messagebox.askyesno("Файлы уже существуют", 
            f"Файлы уже существуют: {', '.join(existing_files_info)}. Заменить их?")

        if replace:
            for _, docx_path, pdf_path in existing_files:
                if os.path.exists(docx_path):
                    os.remove(docx_path)
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
        else:
            # Создаем уникальные имена для файлов и сохраняем их
            for file_name, _, _ in existing_files:
                unique_file_names[file_name] = generate_unique_suffix(file_name)

    for row_data in akt_list:
        context = {}
        for variable, column in header_row.items():
            index = header_dict.get(variable)
            if index is not None:
                context[variable] = row_data[index]
            else:
                context[variable] = ''

        first_row_header = list(header_row.keys())[0]
        file_name = re.sub(r'[\\/*?:"<>|]', '_', str(context[column_name]))

        docx_file_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.docx")
        pdf_file_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.pdf")

        if existing_files and file_name in unique_file_names:
            file_name = unique_file_names[file_name]  # Используем уникальное имя из сохраненных

        if any(context.values()):
            doc = DocxTemplate(TEMPLATE_PATH)
            doc.render(context)
            doc_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.docx")
            doc.save(doc_path)

            if convert_to_pdf:
                pdf_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.pdf")
                convert(doc_path, pdf_path)
                if delete_docx:
                    os.remove(doc_path)
        else:
            pass

    delete_files_with_pattern(SAVE_PATH, f"шаблон-{column_name}")

# Функция для чтения данных из файла Excel
def excel_read(root, path_file):
    if not TEMPLATE_PATH or not EXCEL_PATH or not SAVE_PATH:
        messagebox.showinfo("Ошибка", "Выберите все необходимые файлы и путь")
        return

    wb_read = op.load_workbook(filename=path_file, data_only=True)
    sheet_read = wb_read.active

    header_row = {cell.value: cell.column_letter for cell in sheet_read[1]}
    template_variables = parse_template(TEMPLATE_PATH)

    differences = compare_headers_and_variables(header_row, template_variables)
    if differences != "Заголовки в файле Ecel совпадают с переменными в шаблоне.":
        show_header_and_variable_selection_ui(root, header_row, template_variables)
    else:
        default_column_name = next(iter(header_row.keys()))
        create_doc(root, [], header_row, 1, default_column_name)

# Функция для обработки закрытия окна
def on_closing(root):
    if messagebox.askokcancel("Выход", "Вы уверены, что хотите выйти?"):
        root.destroy()

# Функция для отображения информации о разработчике
def show_developer_info():
    developer_window = Tk()
    developer_window.title("О разработчике")

    developer_label = Label(developer_window, text="Программный продукт был разработан для облегчения Вашей работы", padx=10, pady=5)
    developer_label.pack()
    developer_label = Label(developer_window, text="Разработчик - https://github.com/vok32", padx=10, pady=5)
    developer_label.pack()

    back_button = Button(developer_window, text="Назад", command=developer_window.destroy)
    back_button.pack()

    developer_window.mainloop()

# Основная функция, запускающая программу
if __name__ == '__main__':
    root = Tk()
    root.title("Автозаполнение документов")

    description_label = Label(root, text="Это программа для автозаполнения шаблонов документов Word при помощи таблицы Excel", wraplength=380)
    description_label.grid(row=0, column=0, columnspan=2, padx=10, pady=5)

    select_template_button = Button(root, text="Выбрать файл шаблона Word", command=select_template_file)
    select_template_button.grid(row=1, column=0, padx=10, pady=5, sticky='nsew')
    template_file_label = Label(root, text="", justify="left")
    template_file_label.grid(row=1, column=1, padx=10, pady=5, sticky='nsew')

    select_excel_button = Button(root, text="Выбрать файл с данными Excel", command=select_excel_file)
    select_excel_button.grid(row=2, column=0, padx=10, pady=5, sticky='nsew')
    excel_file_label = Label(root, text="", justify="left")
    excel_file_label.grid(row=2, column=1, padx=10, pady=5, sticky='nsew')

    select_save_button = Button(root, text="Выбрать путь сохранения собранных файлов", command=select_save_folder)
    select_save_button.grid(row=3, column=0, padx=10, pady=5, sticky='nsew')
    save_folder_label = Label(root, text="", justify="left")
    save_folder_label.grid(row=3, column=1, padx=10, pady=5, sticky='nsew')

    start_button = Button(root, text="Продолжить", command=lambda: excel_read(root, EXCEL_PATH))
    start_button.grid(row=4, column=0, padx=10, pady=5, sticky='nsew')

    developer_button = Button(root, text="Об разработчике", command=show_developer_info)
    developer_button.grid(row=5, column=0, padx=10, pady=5, sticky='nsew')

    root.protocol("WM_DELETE_WINDOW", lambda: on_closing(root))
    root.mainloop()
