import re
import os
import sys
import subprocess
import openpyxl as op
from tkinter import Tk, Label, Button, Listbox, Scrollbar, filedialog, messagebox, StringVar, Frame, LEFT
from docx2txt import process
from docxtpl import DocxTemplate
from docx2pdf import convert
import random
import string

# Глобальные переменные для хранения путей к файлам и папкам
TEMPLATE_PATH = ""
EXCEL_PATH = ""
SAVE_PATH = ""

# Функция для закрытия программы
def close_program():
    exit()

# Функция для перезапуска программы
def restart_program():
    python = sys.executable
    os.execl(python, python, *sys.argv)

# Функция для генерации уникального суффикса
def generate_unique_suffix(length=4):
    return ''.join(random.choices(string.ascii_uppercase, k=length))

# Функция для парсинга переменных из шаблона Word
def parse_template(template_path):
    text = process(template_path)
    variables = set(re.findall(r'\{\{(.+?)\}\}', text))  # Используем множество для уникальных переменных
    return variables

# Функция для выбора файла шаблона Word
def select_template_file():
    global TEMPLATE_PATH
    TEMPLATE_PATH = filedialog.askopenfilename(title="Выберите файл шаблона Word", filetypes=[("Word files", "*.docx")])
    template_file_label.config(text=os.path.basename(TEMPLATE_PATH))

# Функция для выбора файла Excel с данными
def select_excel_file():
    global EXCEL_PATH
    EXCEL_PATH = filedialog.askopenfilename(title="Выберите файл с данными Excel", filetypes=[("Excel files", "*.xlsx")])
    excel_file_label.config(text=os.path.basename(EXCEL_PATH))

# Функция для выбора папки для сохранения собранных файлов
def select_save_folder():
    global SAVE_PATH
    SAVE_PATH = filedialog.askdirectory(title="Выберите путь сохранения собранных файлов")
    save_folder_label.config(text=SAVE_PATH)

# Функция для открытия папки с собранными файлами
def open_folder():
    if os.path.exists(SAVE_PATH):
        os.startfile(SAVE_PATH)
    else:
        messagebox.showinfo("Папка не найдена", "Папка с собранными файлами не существует или не была выбрана.")

# Функция для сравнения заголовков Excel и переменных шаблона
def compare_headers_and_variables(header_row, template_variables):
    excel_headers_set = set(header_row.keys())
    template_variables_set = set(template_variables)

    if excel_headers_set != template_variables_set:
        missing_in_excel = excel_headers_set - template_variables_set
        missing_in_template = template_variables_set - excel_headers_set
        message = "Отличия между заголовками в файле Excel и переменными в шаблоне:\n\n"
        if missing_in_excel:
            message += "Заголовки в файле Excel, но отсутствующие в шаблоне:\n"
            for header in missing_in_excel:
                message += f"- {header}\n"
            message += "\n"
        if missing_in_template:
            message += "Переменные в шаблоне, но отсутствующие в заголовках файла Excel:\n"
            for var in missing_in_template:
                message += f"- {var}\n"
        return message
    else:
        return "Заголовки в файле Excel совпадают с переменными в шаблоне."

# Функция для отображения интерфейса выбора заголовков и переменных
def show_header_and_variable_selection_ui(root, header_row, template_variables):
    clear_window(root)

    excel_headers_message = "Список заголовков в файле Excel:\n"
    for index, column_name in enumerate(header_row, start=1):
        excel_headers_message += f"{index}. {column_name}\n"

    template_variables_message = ""
    if template_variables:
        template_variables_message = "Переменные в шаблоне:\n"
        for index, var in enumerate(template_variables, start=1):
            template_variables_message += f"{index}. {var}\n"

    excel_headers_label = Label(root, text=excel_headers_message, justify="left")
    excel_headers_label.grid(row=0, column=0, sticky="w")

    template_variables_label = Label(root, text=template_variables_message, justify="left")
    template_variables_label.grid(row=1, column=0, sticky="w")

    continue_button = Button(root, text="Продолжить", command=lambda: show_differences_ui(root, header_row, template_variables))
    continue_button.grid(row=2, column=0, pady=10)

    close_button = Button(root, text="Закрыть программу", command=close_program)
    close_button.grid(row=4, column=0, pady=5)

# Функция для отображения интерфейса различий между заголовками и переменными
def show_differences_ui(root, header_row, template_variables):
    clear_window(root)

    differences_message = compare_headers_and_variables(header_row, template_variables)
    differences_label = Label(root, text=differences_message, justify="left")
    differences_label.grid(row=0, column=0, sticky="w")

    max_width = max(len(line) for line in differences_message.split('\n'))
    set_label_width(differences_label, max_width)

    replace_button = Button(root, text="Продолжить и заменить", command=lambda: select_column(root, header_row, template_variables))
    replace_button.grid(row=1, column=0, pady=5)

    close_button = Button(root, text="Закрыть программу", command=close_program)
    close_button.grid(row=3, column=0, pady=5)

# Функция для установки ширины метки
def set_label_width(label, max_width):
    label.config(width=max_width)

# Функция для выбора столбца Excel
def select_column(root, header_row, template_variables):
    clear_window(root)

    wb_read = op.load_workbook(filename=EXCEL_PATH, data_only=True)
    sheet_read = wb_read.active

    akt_list = [list(row) for row in sheet_read.iter_rows(values_only=True)]

    def select_column_callback():
        selected_index = listbox.curselection()
        if selected_index:
            selected_column_name = listbox.get(selected_index[0])
            selected_column = header_row[selected_column_name]
            choice_root = Tk()
            choice_root.title("Выбор формата файла")

            def docx_pdf():
                choice_root.destroy()
                create_doc(root, akt_list, header_row, selected_column, selected_column_name, convert_to_pdf=True, delete_docx=False)
                show_success_or_report_window(root)

            def docx_only():
                choice_root.destroy()
                create_doc(root, akt_list, header_row, selected_column, selected_column_name, convert_to_pdf=False)
                show_success_or_report_window(root)

            def pdf_only():
                choice_root.destroy()
                create_doc(root, akt_list, header_row, selected_column, selected_column_name, convert_to_pdf=True, delete_docx=True)
                show_success_or_report_window(root)

            pdf_button = Button(choice_root, text="DOCX and PDF", command=docx_pdf)
            pdf_button.grid(row=0, column=0, padx=10, pady=5)

            docx_button = Button(choice_root, text="DOCX Only", command=docx_only)
            docx_button.grid(row=0, column=1, padx=10, pady=5)

            pdf_only_button = Button(choice_root, text="PDF Only", command=pdf_only)
            pdf_only_button.grid(row=1, column=0, columnspan=2, padx=10, pady=5)

            choice_root.mainloop()
        else:
            messagebox.showinfo("Ошибка", "Не выбрано название для файла. Пожалуйста, выберите название.")

    selection_label = Label(root, text="Выберите название для файлов:")
    selection_label.grid(row=0, column=0, sticky="w")

    listbox = Listbox(root, selectmode="SINGLE")
    listbox.grid(row=1, column=0, sticky="w")

    scrollbar = Scrollbar(root, orient="vertical")
    scrollbar.config(command=listbox.yview)
    scrollbar.grid(row=1, column=1, sticky="ns")

    listbox.config(yscrollcommand=scrollbar.set)
    for column_name in header_row:
        listbox.insert("end", column_name)

    select_button = Button(root, text="Выбрать", command=select_column_callback)
    select_button.grid(row=2, column=0, pady=10)

    close_button = Button(root, text="Закрыть программу", command=close_program)
    close_button.grid(row=4, column=0, pady=5)

# Функция для отображения окна успешного выполнения или отчёта об ошибке
def show_success_or_report_window(root):
    clear_window(root)

    success_label = Label(root, text="Замена прошла успешно", justify="left", fg="green", font=("Arial", 12))
    success_label.grid(row=0, column=0, pady=5)

    open_folder_button = Button(root, text="Открыть папку с собранными файлами", command=open_folder)
    open_folder_button.grid(row=1, column=0, pady=5)

    report_label = Label(root, text="Вы хотите отправить разработчику отчёт об ошибке в приложении?", justify="left")
    report_label.grid(row=2, column=0, pady=5)

    yes_button = Button(root, text="Да, отправить", command=lambda: messagebox.showinfo("Ну и ябеда!", "Ну и ябеда!"))
    yes_button.grid(row=3, column=0, pady=5)

    buttons_frame = Frame(root)
    buttons_frame.grid(row=4, column=0, pady=5)

    restart_button = Button(buttons_frame, text="Запустить программу снова", command=restart_program)
    restart_button.pack(side=LEFT, padx=5)

    close_button = Button(buttons_frame, text="Закрыть программу", command=close_program)
    close_button.pack(side=LEFT, padx=5)

# Функция для очистки основного окна
def clear_window(root):
    for widget in root.winfo_children():
        widget.destroy()

# Функция для создания документа
def create_doc(root, akt_list, header_row, column_index, column_name, convert_to_pdf=True, delete_docx=True):
    header_dict = {name: index for index, name in enumerate(header_row)}

    for row_data in akt_list:
        context = {}
        for variable, column in header_row.items():
            index = header_dict.get(variable)
            if index is not None:
                context[variable] = row_data[index]
            else:
                context[variable] = ''

        first_row_header = list(header_row.keys())[0]
        file_name = re.sub(r'[\\/*?:"<>|]', '_', str(context[column_name]))

        file_exists = os.path.exists(os.path.join(SAVE_PATH, f"шаблон-{file_name}.docx"))
        if file_exists:
            file_name += f"_{generate_unique_suffix()}"

        if any(context.values()):
            doc = DocxTemplate(TEMPLATE_PATH)
            doc.render(context)
            doc_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.docx")
            doc.save(doc_path)

            if convert_to_pdf:
                pdf_path = os.path.join(SAVE_PATH, f"шаблон-{file_name}.pdf")
                convert(doc_path, pdf_path)
                if delete_docx:
                    os.remove(doc_path)
        else:
            pass

    selected_file_path = os.path.join(SAVE_PATH, f"шаблон-{column_name}.docx")
    if os.path.exists(selected_file_path):
        os.remove(selected_file_path)

    selected_file_pdf_path = os.path.join(SAVE_PATH, f"шаблон-{column_name}.pdf")
    if os.path.exists(selected_file_pdf_path):
        os.remove(selected_file_pdf_path)

# Функция для чтения данных из файла Excel
def excel_read(root, path_file):
    if not TEMPLATE_PATH or not EXCEL_PATH or not SAVE_PATH:
        messagebox.showinfo("Ошибка", "Выберите все необходимые файлы и путь")
        return

    wb_read = op.load_workbook(filename=path_file, data_only=True)
    sheet_read = wb_read.active

    header_row = {cell.value: cell.column_letter for cell in sheet_read[1]}
    template_variables = parse_template(TEMPLATE_PATH)

    differences = compare_headers_and_variables(header_row, template_variables)
    if differences != "Заголовки в файле Ecel совпадают с переменными в шаблоне.":
        show_header_and_variable_selection_ui(root, header_row, template_variables)
    else:
        default_column_name = next(iter(header_row.keys()))
        create_doc(root, [], header_row, 1, default_column_name)

# Функция для обработки закрытия окна
def on_closing(root):
    if messagebox.askokcancel("Выход", "Вы уверены, что хотите выйти?"):
        root.destroy()

# Функция для отображения информации о разработчике
def show_developer_info():
    developer_window = Tk()
    developer_window.title("О разработчике")

    developer_label = Label(developer_window, text="Программный продукт был разработан для облегчения Вашей работы", padx=10, pady=5)
    developer_label.pack()
    developer_label = Label(developer_window, text="Разработчик - https://github.com/vok32", padx=10, pady=5)
    developer_label.pack()

    back_button = Button(developer_window, text="Назад", command=developer_window.destroy)
    back_button.pack()

    developer_window.mainloop()

# Основная функция, запускающая программу
if __name__ == '__main__':
    root = Tk()
    root.title("Автозаполнение документов")

    description_label = Label(root, text="Это программа для автозаполнения шаблонов документов Word при помощи таблицы Excel", wraplength=380)
    description_label.grid(row=0, column=0, columnspan=2, padx=10, pady=5)

    select_template_button = Button(root, text="Выбрать файл шаблона Word", command=select_template_file)
    select_template_button.grid(row=1, column=0, padx=10, pady=5, sticky='nsew')
    template_file_label = Label(root, text="", justify="left")
    template_file_label.grid(row=1, column=1, padx=10, pady=5, sticky='nsew')

    select_excel_button = Button(root, text="Выбрать файл с данными Excel", command=select_excel_file)
    select_excel_button.grid(row=2, column=0, padx=10, pady=5, sticky='nsew')
    excel_file_label = Label(root, text="", justify="left")
    excel_file_label.grid(row=2, column=1, padx=10, pady=5, sticky='nsew')

    select_save_button = Button(root, text="Выбрать путь сохранения собранных файлов", command=select_save_folder)
    select_save_button.grid(row=3, column=0, padx=10, pady=5, sticky='nsew')
    save_folder_label = Label(root, text="", justify="left")
    save_folder_label.grid(row=3, column=1, padx=10, pady=5, sticky='nsew')

    start_button = Button(root, text="Продолжить", command=lambda: excel_read(root, EXCEL_PATH))
    start_button.grid(row=4, column=0, padx=10, pady=5, sticky='nsew')

    developer_button = Button(root, text="Об разработчике", command=show_developer_info)
    developer_button.grid(row=5, column=0, padx=10, pady=5, sticky='nsew')

    root.protocol("WM_DELETE_WINDOW", lambda: on_closing(root))
    root.mainloop()
